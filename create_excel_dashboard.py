import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
EXCEL_DIR = os.path.join(BASE_DIR, "excel")

os.makedirs(EXCEL_DIR, exist_ok=True)

raw_path = os.path.join(DATA_DIR, "bindisa_agriculture_yearly_data.csv")
output_path = os.path.join(EXCEL_DIR, "Bindisa_Agriculture_Advanced_Dashboard.xlsx")

df = pd.read_csv(raw_path)

yearly = (
    df.groupby("Year")
    .agg(
        Total_Revenue=("Revenue", "sum"),
        Total_Cost=("Total_Cost", "sum"),
        Total_Profit=("Profit", "sum"),
        Total_Production=("Production_Tonnes", "sum"),
        Average_Yield=("Yield_Tonnes_Per_Hectare", "mean"),
    )
    .reset_index()
)

crop = (
    df.groupby("Crop")
    .agg(Total_Revenue=("Revenue", "sum"), Total_Profit=("Profit", "sum"))
    .reset_index()
    .sort_values("Total_Profit", ascending=False)
)

region = (
    df.groupby("Region")
    .agg(Total_Revenue=("Revenue", "sum"), Total_Profit=("Profit", "sum"))
    .reset_index()
    .sort_values("Total_Revenue", ascending=False)
)

cost = df[
    [
        "Fertilizer_Cost",
        "Pesticide_Cost",
        "Irrigation_Cost",
        "Labor_Cost",
        "Storage_Transport_Cost",
    ]
].sum().reset_index()
cost.columns = ["Cost_Category", "Amount"]

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Raw_Data", index=False)
    yearly.to_excel(writer, sheet_name="Yearly_Summary", index=False)
    crop.to_excel(writer, sheet_name="Crop_Summary", index=False)
    region.to_excel(writer, sheet_name="Region_Summary", index=False)
    cost.to_excel(writer, sheet_name="Cost_Summary", index=False)

wb = load_workbook(output_path)
ws = wb.create_sheet("Dashboard", 0)

title_fill = PatternFill("solid", fgColor="1F4E78")
kpi_fill = PatternFill("solid", fgColor="D9EAF7")
header_fill = PatternFill("solid", fgColor="70AD47")
white_font = Font(color="FFFFFF", bold=True, size=14)
header_font = Font(color="FFFFFF", bold=True)

ws["A1"] = "Bindisa Agriculture Private Limited"
ws["A2"] = "Yearly Agricultural Data Analytics and Business Intelligence Dashboard"
ws["A1"].font = Font(bold=True, size=20, color="1F4E78")
ws["A2"].font = Font(bold=True, size=12, color="548235")

total_revenue = df["Revenue"].sum()
total_profit = df["Profit"].sum()
total_cost = df["Total_Cost"].sum()
total_production = df["Production_Tonnes"].sum()
avg_yield = df["Yield_Tonnes_Per_Hectare"].mean()
profit_margin = total_profit / total_revenue * 100

kpis = [
    ("Total Revenue", total_revenue),
    ("Total Profit", total_profit),
    ("Total Cost", total_cost),
    ("Total Production", total_production),
    ("Average Yield", avg_yield),
    ("Profit Margin %", profit_margin),
]

start_col = 1
for index, (label, value) in enumerate(kpis):
    col = start_col + index * 2
    ws.cell(row=4, column=col).value = label
    ws.cell(row=4, column=col).fill = title_fill
    ws.cell(row=4, column=col).font = white_font
    ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")

    ws.cell(row=5, column=col).value = round(value, 2)
    ws.cell(row=5, column=col).fill = kpi_fill
    ws.cell(row=5, column=col).font = Font(bold=True, size=12)
    ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")

for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max_length + 3, 35)

yearly_sheet = wb["Yearly_Summary"]
crop_sheet = wb["Crop_Summary"]
cost_sheet = wb["Cost_Summary"]

line_chart = LineChart()
line_chart.title = "Yearly Revenue Trend"
line_chart.y_axis.title = "Revenue"
line_chart.x_axis.title = "Year"
line_chart.add_data(Reference(yearly_sheet, min_col=2, min_row=1, max_row=yearly_sheet.max_row), titles_from_data=True)
line_chart.set_categories(Reference(yearly_sheet, min_col=1, min_row=2, max_row=yearly_sheet.max_row))
ws.add_chart(line_chart, "A8")

bar_chart = BarChart()
bar_chart.title = "Crop-wise Profit"
bar_chart.y_axis.title = "Profit"
bar_chart.x_axis.title = "Crop"
bar_chart.add_data(Reference(crop_sheet, min_col=3, min_row=1, max_row=crop_sheet.max_row), titles_from_data=True)
bar_chart.set_categories(Reference(crop_sheet, min_col=1, min_row=2, max_row=crop_sheet.max_row))
ws.add_chart(bar_chart, "I8")

pie_chart = PieChart()
pie_chart.title = "Cost Distribution"
pie_chart.add_data(Reference(cost_sheet, min_col=2, min_row=1, max_row=cost_sheet.max_row), titles_from_data=True)
pie_chart.set_categories(Reference(cost_sheet, min_col=1, min_row=2, max_row=cost_sheet.max_row))
ws.add_chart(pie_chart, "A24")

wb.save(output_path)
print(f"Excel dashboard created successfully: {output_path}")

